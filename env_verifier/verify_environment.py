#!/usr/bin/env python3
"""
verify_environment.py

Lightweight environment verification for the AI for NTH Summer School.

What it checks:
1. Python version
2. Imports of required packages
3. Basic filesystem layout (optional / non-fatal)
4. Small numerical smoke tests:
   - NumPy / pandas
   - matplotlib plot save
   - scikit-learn linear regression
   - scikit-learn random forest
   - scikit-learn Gaussian process regression
   - openpyxl workbook write
   - PyTorch CPU tensor / tiny model forward pass
5. JupyterLab availability

Intended use:
    python verify_environment.py

Exit status:
    0 = all critical checks passed
    1 = one or more critical checks failed
"""

from __future__ import annotations

import importlib
import json
import os
import platform
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, List, Tuple


MIN_PYTHON = (3, 11)

REQUIRED_MODULES = {
    "numpy": "numpy",
    "pandas": "pandas",
    "scipy": "scipy",
    "matplotlib": "matplotlib",
    "scikit-learn": "sklearn",
    "joblib": "joblib",
    "openpyxl": "openpyxl",
    "pyyaml": "yaml",
    "torch": "torch",
    "jupyterlab": "jupyterlab",
}


@dataclass
class CheckResult:
    name: str
    passed: bool
    detail: str


def print_header(title: str) -> None:
    line = "=" * len(title)
    print(f"\n{title}\n{line}")


def format_bool(flag: bool) -> str:
    return "PASS" if flag else "FAIL"


def module_version(mod: Any) -> str:
    return getattr(mod, "__version__", "unknown")


def get_bundle_root() -> Path:
    """
    Tries to infer the summer school bundle root.
    Expected bundle layout:
      bundle/
        notebooks/
        data/
        verify/
          verify_environment.py

    Fallback is the current working directory.
    """
    here = Path(__file__).resolve()
    candidates = [
        here.parent.parent,  # if script is in verify/
        here.parent,         # if script is in bundle root
        Path.cwd(),
    ]
    for cand in candidates:
        if (cand / "notebooks").exists() or (cand / "data").exists() or (cand / "verify").exists():
            return cand
    return Path.cwd()


def check_python() -> CheckResult:
    ok = sys.version_info >= MIN_PYTHON
    detail = (
        f"Python executable: {sys.executable}\n"
        f"Python version: {platform.python_version()} "
        f"(required >= {MIN_PYTHON[0]}.{MIN_PYTHON[1]})"
    )
    return CheckResult("Python version", ok, detail)


def check_imports() -> Tuple[List[CheckResult], Dict[str, Any]]:
    results: List[CheckResult] = []
    imported: Dict[str, Any] = {}
    for label, module_name in REQUIRED_MODULES.items():
        try:
            mod = importlib.import_module(module_name)
            imported[module_name] = mod
            results.append(
                CheckResult(
                    f"Import: {label}",
                    True,
                    f"Imported '{module_name}' successfully (version: {module_version(mod)})",
                )
            )
        except Exception as exc:  # pragma: no cover - broad by design for verification
            results.append(
                CheckResult(
                    f"Import: {label}",
                    False,
                    f"Failed to import '{module_name}': {exc}",
                )
            )
    return results, imported


def check_filesystem(bundle_root: Path) -> List[CheckResult]:
    expected = {
        "Bundle root": bundle_root,
        "notebooks/": bundle_root / "notebooks",
        "data/": bundle_root / "data",
        "verify/": bundle_root / "verify",
    }
    results: List[CheckResult] = []
    for label, path in expected.items():
        exists = path.exists()
        results.append(CheckResult(f"Filesystem: {label}", exists, f"Path: {path}"))
    return results


def smoke_test_numpy_pandas(np_mod: Any, pd_mod: Any) -> CheckResult:
    try:
        arr = np_mod.array([[1.0, 2.0], [3.0, 4.0]])
        df = pd_mod.DataFrame(arr, columns=["a", "b"])
        mean_a = float(df["a"].mean())
        ok = abs(mean_a - 2.0) < 1e-12
        return CheckResult(
            "Smoke test: NumPy + pandas",
            ok,
            f"Created array/DataFrame successfully; mean(a)={mean_a}",
        )
    except Exception as exc:
        return CheckResult("Smoke test: NumPy + pandas", False, str(exc))


def smoke_test_matplotlib(matplotlib_mod: Any) -> CheckResult:
    try:
        matplotlib_mod.use("Agg", force=True)
        import matplotlib.pyplot as plt  # imported after backend set

        with tempfile.TemporaryDirectory() as tmpdir:
            out = Path(tmpdir) / "smoke_plot.png"
            fig = plt.figure()
            ax = fig.add_subplot(111)
            ax.plot([0, 1, 2], [0, 1, 4])
            ax.set_title("Environment verification plot")
            fig.tight_layout()
            fig.savefig(out)
            plt.close(fig)
            ok = out.exists() and out.stat().st_size > 0
        return CheckResult(
            "Smoke test: matplotlib",
            ok,
            "Saved a simple plot successfully",
        )
    except Exception as exc:
        return CheckResult("Smoke test: matplotlib", False, str(exc))


def smoke_test_sklearn(imported: Dict[str, Any]) -> List[CheckResult]:
    results: List[CheckResult] = []
    try:
        import numpy as np
        from sklearn.ensemble import RandomForestRegressor
        from sklearn.gaussian_process import GaussianProcessRegressor
        from sklearn.gaussian_process.kernels import RBF, WhiteKernel
        from sklearn.linear_model import LinearRegression

        X = np.linspace(0.0, 1.0, 20).reshape(-1, 1)
        y = (X[:, 0] ** 2 + 0.1).astype(float)

        lin = LinearRegression().fit(X, y)
        lin_pred = float(lin.predict([[0.5]])[0])

        rf = RandomForestRegressor(
            n_estimators=10,
            max_depth=3,
            random_state=0,
            n_jobs=1,
        ).fit(X, y)
        rf_pred = float(rf.predict([[0.5]])[0])

        kernel = RBF(length_scale=0.2) + WhiteKernel(noise_level=1e-4)
        gp = GaussianProcessRegressor(kernel=kernel, random_state=0).fit(X, y)
        gp_mean, gp_std = gp.predict([[0.5]], return_std=True)
        gp_mean = float(gp_mean[0])
        gp_std = float(gp_std[0])

        results.append(
            CheckResult(
                "Smoke test: scikit-learn linear regression",
                True,
                f"LinearRegression fit/predict OK; prediction at x=0.5 is {lin_pred:.6f}",
            )
        )
        results.append(
            CheckResult(
                "Smoke test: scikit-learn random forest",
                True,
                f"RandomForestRegressor fit/predict OK; prediction at x=0.5 is {rf_pred:.6f}",
            )
        )
        results.append(
            CheckResult(
                "Smoke test: scikit-learn Gaussian process",
                True,
                f"GPR fit/predict OK; mean={gp_mean:.6f}, std={gp_std:.6f}",
            )
        )
    except Exception as exc:
        results.append(CheckResult("Smoke test: scikit-learn models", False, str(exc)))
    return results


def smoke_test_openpyxl() -> CheckResult:
    try:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            out = Path(tmpdir) / "openpyxl_test.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "AI for NTH"
            ws["A2"] = 42
            wb.save(out)
            ok = out.exists() and out.stat().st_size > 0
        return CheckResult("Smoke test: openpyxl", ok, "Workbook created and saved successfully")
    except Exception as exc:
        return CheckResult("Smoke test: openpyxl", False, str(exc))


def smoke_test_torch() -> CheckResult:
    try:
        import torch
        from torch import nn

        x = torch.tensor([[1.0, 2.0], [3.0, 4.0]])
        model = nn.Sequential(nn.Linear(2, 4), nn.ReLU(), nn.Linear(4, 1))
        with torch.no_grad():
            y = model(x)
        ok = tuple(y.shape) == (2, 1)
        detail = (
            f"PyTorch version {torch.__version__}; "
            f"CUDA available: {torch.cuda.is_available()}; "
            f"tiny forward pass output shape: {tuple(y.shape)}"
        )
        return CheckResult("Smoke test: PyTorch CPU", ok, detail)
    except Exception as exc:
        return CheckResult("Smoke test: PyTorch CPU", False, str(exc))


def check_jupyterlab() -> CheckResult:
    """
    Prefer checking via the current Python executable so we verify the active environment.
    """
    try:
        completed = subprocess.run(
            [sys.executable, "-m", "jupyter", "lab", "--version"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=True,
        )
        version = completed.stdout.strip() or completed.stderr.strip() or "unknown"
        return CheckResult("JupyterLab availability", True, f"jupyter lab version: {version}")
    except Exception as exc:
        return CheckResult("JupyterLab availability", False, str(exc))


def summarize(results: List[CheckResult]) -> Tuple[int, int]:
    passed = sum(1 for r in results if r.passed)
    total = len(results)
    return passed, total


def print_results(results: List[CheckResult]) -> None:
    for result in results:
        print(f"[{format_bool(result.passed)}] {result.name}")
        print(f"      {result.detail}")


def main() -> int:
    all_results: List[CheckResult] = []

    print_header("AI for NTH Summer School - Environment Verification")

    # System info
    bundle_root = get_bundle_root()
    print(f"Bundle root guess: {bundle_root}")
    print(f"Platform: {platform.platform()}")
    print(f"Machine: {platform.machine()}")
    print(f"Working directory: {Path.cwd()}")
    print(f"Script location: {Path(__file__).resolve()}")

    # Python
    print_header("1. Python")
    py_result = check_python()
    all_results.append(py_result)
    print_results([py_result])

    # Imports
    print_header("2. Imports")
    import_results, imported = check_imports()
    all_results.extend(import_results)
    print_results(import_results)

    # Filesystem (non-critical, but useful)
    print_header("3. Bundle / filesystem layout")
    fs_results = check_filesystem(bundle_root)
    all_results.extend(fs_results)
    print_results(fs_results)

    # Smoke tests
    print_header("4. Smoke tests")
    if imported.get("numpy") and imported.get("pandas"):
        all_results.append(smoke_test_numpy_pandas(imported["numpy"], imported["pandas"]))
    if imported.get("matplotlib"):
        all_results.append(smoke_test_matplotlib(imported["matplotlib"]))
    if imported.get("sklearn"):
        all_results.extend(smoke_test_sklearn(imported))
    if imported.get("openpyxl"):
        all_results.append(smoke_test_openpyxl())
    if imported.get("torch"):
        all_results.append(smoke_test_torch())

    smoke_results = [r for r in all_results if r.name.startswith("Smoke test:")]
    print_results(smoke_results)

    # Jupyter
    print_header("5. JupyterLab")
    jup_result = check_jupyterlab()
    all_results.append(jup_result)
    print_results([jup_result])

    # Summary
    print_header("Summary")
    passed, total = summarize(all_results)
    critical_failures = [
        r for r in all_results
        if (not r.passed) and not r.name.startswith("Filesystem:")
    ]

    print(f"Checks passed: {passed} / {total}")
    if critical_failures:
        print("\nCritical failures detected:")
        for r in critical_failures:
            print(f" - {r.name}: {r.detail}")
        print("\nEnvironment verification FAILED.")
        return 1

    print("\nEnvironment verification PASSED.")
    print("You should now be able to run the notebooks for the summer school.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
